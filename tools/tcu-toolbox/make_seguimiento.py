#!/usr/bin/env python3
"""Genera la ficha Excel "Seguimiento PEM" de una planta desde su fichero de
plantas (JSON o CSV de la toolbox / plataforma).

    python make_seguimiento.py plantas/elburgo.json [-o Seguimiento_PEM_ElBurgo.xlsx]

Estructura (misma que la ficha de El Burgo):
  - Hoja "Resumen": una fila por NCU con formulas que cuentan el avance de las
    pestanas de NCU (no se rellena a mano) + fila TOTAL + barra de % avance.
  - Una hoja por NCU: configuracion de NCU, hasta 4 HSUs (una fila por HSU con
    nombre; sin nombre no cuentan), y una fila por TCU con sus tres tareas
    (Cold commissioning / Configuracion TCU / Prueba movimiento), desplegable
    OK / NOK / N.A., fecha, tecnico y observaciones. Amarillo = rellenar.

Requiere: openpyxl (pip install openpyxl).
"""
import argparse
import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- estilos (los de la ficha original de El Burgo) ---
AZUL_OSCURO = '1F4E79'   # cabecera del Resumen
AZUL_MEDIO = '2E75B6'    # cabeceras de las hojas de NCU
AMARILLO = 'FFF2CC'      # celdas a rellenar
VERDE_F, VERDE_T = 'C6EFCE', '006100'   # OK
ROJO_F, ROJO_T = 'FFC7CE', '9C0006'     # NOK
FINO = Side(style='thin', color='BFBFBF')
BORDE = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)
CENTRO = Alignment(horizontal='center')

RE_NCU = re.compile(r'^(.*?)\s+NCU(\d+)\b', re.IGNORECASE)
MAX_HSUS = 4          # filas de HSU por NCU en la ficha
FILA_TCU1 = 13        # primera fila de la tabla de TCUs
FONDO_FILAS = 120     # las formulas cubren hasta aqui (o mas si hay mas TCUs)


def f(ws, coord, valor, size=10, bold=False, fill=None, color=None, border=False,
      center=False, numfmt=None):
    c = ws[coord]
    c.value = valor
    c.font = Font(name='Arial', size=size, bold=bold, color=color)
    if fill:
        c.fill = PatternFill('solid', start_color=fill)
    if border:
        c.border = BORDE
    if center:
        c.alignment = CENTRO
    if numfmt:
        c.number_format = numfmt
    return c


def cargar_topologia(ruta):
    """Devuelve (nombre_planta, {ncu: {'ip': .., 'gws': [(puerto, ini, fin)...], 'hsu': ..}})."""
    entradas = []
    if ruta.suffix.lower() == '.json':
        datos = json.loads(ruta.read_text(encoding='utf-8-sig'))
        for e in datos.get('plantas', []):
            entradas.append({'nombre': str(e.get('nombre', '')), 'ip': str(e.get('ip', '')),
                             'puerto': int(e.get('puerto', 0)), 'ini': int(e.get('tcu_ini', 0)),
                             'fin': int(e.get('tcu_fin', 0)), 'hsu': e.get('hsu_esclavo')})
    else:
        with ruta.open(encoding='utf-8-sig', newline='') as fh:
            for fila in csv.reader(fh, delimiter=';'):
                if len(fila) < 6 or fila[0].strip().lower() == 'planta':
                    continue
                planta, ncu, ip, puerto, ini, fin = (x.strip() for x in fila[:6])
                entradas.append({'nombre': f'{planta} {ncu}', 'ip': ip, 'puerto': int(puerto),
                                 'ini': int(ini), 'fin': int(fin), 'hsu': None})
    ncus = {}
    nombre_planta = None
    for e in entradas:
        m = RE_NCU.match(e['nombre'])
        if not m:
            continue
        nombre_planta = nombre_planta or m.group(1).strip()
        n = int(m.group(2))
        d = ncus.setdefault(n, {'ip': e['ip'], 'gws': [], 'hsu': None})
        d['gws'].append((e['puerto'], e['ini'], e['fin']))
        if e['hsu'] and not d['hsu']:
            d['hsu'] = e['hsu']
    if not ncus:
        raise SystemExit(f'ERROR: {ruta} no tiene entradas "<Planta> NCU<n> ..." reconocibles')
    for d in ncus.values():
        d['gws'].sort(key=lambda g: (g[1], g[0]))
    return nombre_planta or ruta.stem, dict(sorted(ncus.items()))


def hoja_ncu(wb, n, d, con_ejemplo):
    ws = wb.create_sheet(f'NCU{n}')
    tcus = []
    for puerto, ini, fin in d['gws']:
        tcus += [(t, puerto) for t in range(ini, fin + 1)]
    tcus = sorted(set(tcus))
    ult = max(FONDO_FILAS, FILA_TCU1 - 1 + len(tcus))

    f(ws, 'A1', f"Seguimiento PEM - NCU{n}  ({d['ip']})", size=13, bold=True)
    f(ws, 'A2', 'Rellenar solo celdas amarillas (desplegable OK / NOK / N.A.). '
                'HSUs: escribe el nombre en la columna Equipo; las filas sin nombre no cuentan.',
      color='808080')

    for col, tit in zip('ABCD', ['Equipo', 'Tarea', 'Estado', 'Observaciones']):
        f(ws, f'{col}4', tit, bold=True, fill=AZUL_MEDIO, color='FFFFFF', border=True, center=True)
    f(ws, 'A5', f'NCU{n}', border=True)
    f(ws, 'B5', 'Configuracion NCU', border=True)
    f(ws, 'C5', None, fill=AMARILLO, border=True, center=True)
    f(ws, 'D5', None, fill=AMARILLO, border=True)
    for i in range(MAX_HSUS):
        fila = 6 + i
        nombre_hsu = 'HSU1' if (i == 0 and d['hsu']) else None
        f(ws, f'A{fila}', nombre_hsu, fill=AMARILLO, border=True)
        f(ws, f'B{fila}', 'Configuracion HSUs', border=True)
        f(ws, f'C{fila}', None, fill=AMARILLO, border=True, center=True)
        f(ws, f'D{fila}', None, fill=AMARILLO, border=True)

    cab = ['TCU', 'GW', 'Cold commissioning', 'Configuracion TCU', 'Prueba movimiento',
           '% Avance', 'Fecha', 'Tecnico', 'Observaciones']
    for col, tit in zip('ABCDEFGHI', cab):
        f(ws, f'{col}12', tit, bold=True, fill=AZUL_MEDIO, color='FFFFFF', border=True, center=True)
    for k, (tcu, puerto) in enumerate(tcus):
        fila = FILA_TCU1 + k
        f(ws, f'A{fila}', tcu, center=True)
        f(ws, f'B{fila}', puerto, center=True)
        for col in 'CDE':
            f(ws, f'{col}{fila}', None, fill=AMARILLO, border=True, center=True)
        f(ws, f'F{fila}', f'=(COUNTIF(C{fila}:E{fila},"OK")+COUNTIF(C{fila}:E{fila},"N.A."))/3',
          bold=True, center=True, numfmt='0%')
        f(ws, f'G{fila}', None, fill=AMARILLO, border=True, numfmt='dd/mm/yyyy')
        f(ws, f'H{fila}', None, fill=AMARILLO, border=True)
        f(ws, f'I{fila}', None, fill=AMARILLO, border=True)
    if con_ejemplo and tcus:
        for col, v in zip('CDEGHI', ['OK', 'OK', 'NOK', '15/07/2026', 'IMG',
                                     'FILA DE EJEMPLO - sobrescribir con lo real']):
            ws[f'{col}{FILA_TCU1}'] = v

    dv = DataValidation(type='list', formula1='"OK,NOK,N.A."', allow_blank=True)
    dv.add(f'C5:C{5 + MAX_HSUS}')
    dv.add(f'C{FILA_TCU1}:E{ult}')
    ws.add_data_validation(dv)
    for rango in (f'C5:C{5 + MAX_HSUS}', f'C{FILA_TCU1}:E{ult}'):
        ws.conditional_formatting.add(rango, CellIsRule(
            operator='equal', formula=['"OK"'], fill=PatternFill('solid', start_color=VERDE_F),
            font=Font(color=VERDE_T)))
        ws.conditional_formatting.add(rango, CellIsRule(
            operator='equal', formula=['"NOK"'], fill=PatternFill('solid', start_color=ROJO_F),
            font=Font(color=ROJO_T)))
    ws.conditional_formatting.add(f'F{FILA_TCU1}:F{ult}', DataBarRule(
        start_type='num', start_value=0, end_type='num', end_value=1, color='63BE7B'))

    ws.freeze_panes = 'C13'
    for col, ancho in zip('ABCFGHI', [8, 18, 16, 10, 11, 12, 34]):
        ws.column_dimensions[col].width = ancho
    for col in 'DE':
        ws.column_dimensions[col].width = 16
    return ult


def hoja_resumen(wb, planta, ncus, ult_por_ncu):
    ws = wb['Resumen']
    f(ws, 'A1', f'Seguimiento PEM - {planta}', size=15, bold=True)
    f(ws, 'A2', 'Se calcula solo desde las pestanas de NCU. No rellenar aqui.', color='808080')
    cab = ['NCU', 'IP', 'TCUs', 'Config NCU', 'HSUs OK', 'HSUs tot', 'Cold comm.',
           'Config TCU', 'Prueba mov.', 'TCUs 100%', 'Hechos', 'Totales', '% Avance']
    for i, tit in enumerate(cab):
        f(ws, f'{get_column_letter(i + 1)}4', tit, bold=True, fill=AZUL_OSCURO,
          color='FFFFFF', border=True, center=True)
    fila = 5
    for n, d in ncus.items():
        h = f'NCU{n}'
        u = ult_por_ncu[n]
        hh = 5 + MAX_HSUS
        vals = [
            (f'A{fila}', h), (f'B{fila}', d['ip']),
            (f'C{fila}', f'=COUNT({h}!A{FILA_TCU1}:A{u})'),
            (f'D{fila}', f'={h}!C5'),
            (f'E{fila}', f'=COUNTIFS({h}!A6:A{hh},"<>",{h}!C6:C{hh},"OK")'
                         f'+COUNTIFS({h}!A6:A{hh},"<>",{h}!C6:C{hh},"N.A.")'),
            (f'F{fila}', f'=COUNTIF({h}!A6:A{hh},"?*")'),
            (f'G{fila}', f'=COUNTIF({h}!C{FILA_TCU1}:C{u},"OK")+COUNTIF({h}!C{FILA_TCU1}:C{u},"N.A.")'),
            (f'H{fila}', f'=COUNTIF({h}!D{FILA_TCU1}:D{u},"OK")+COUNTIF({h}!D{FILA_TCU1}:D{u},"N.A.")'),
            (f'I{fila}', f'=COUNTIF({h}!E{FILA_TCU1}:E{u},"OK")+COUNTIF({h}!E{FILA_TCU1}:E{u},"N.A.")'),
            (f'J{fila}', f'=COUNTIF({h}!F{FILA_TCU1}:F{u},1)'),
            (f'K{fila}', f'=IF(OR(D{fila}="OK",D{fila}="N.A."),1,0)+E{fila}+G{fila}+H{fila}+I{fila}'),
            (f'L{fila}', f'=1+F{fila}+3*C{fila}'),
            (f'M{fila}', f'=K{fila}/L{fila}'),
        ]
        for coord, v in vals:
            numfmt = '0.0%' if coord.startswith('M') else None
            f(ws, coord, v, border=True, center=(coord[0] != 'B'), bold=coord.startswith('M'),
              numfmt=numfmt)
        fila += 1
    tot = fila
    f(ws, f'A{tot}', 'TOTAL', bold=True, border=True)
    f(ws, f'B{tot}', None, border=True)
    for col in 'CEFGHIJKL':
        f(ws, f'{col}{tot}', f'=SUM({col}5:{col}{tot - 1})', bold=True, border=True, center=True)
    f(ws, f'D{tot}', None, border=True)
    f(ws, f'M{tot}', f'=K{tot}/L{tot}', bold=True, border=True, center=True, numfmt='0.0%')
    ws.conditional_formatting.add(f'M5:M{tot}', DataBarRule(
        start_type='num', start_value=0, end_type='num', end_value=1, color='63BE7B'))
    f(ws, f'A{tot + 2}', 'Tareas: NCU = Configuracion NCU. HSU = Configuracion HSUs (una fila por '
                         'HSU en su pestana). TCU = Cold commissioning, Configuracion TCU, Prueba movimiento.',
      color='808080')
    f(ws, f'A{tot + 3}', 'Hechos/Totales cuentan cada tarea individual; % Avance = Hechos / Totales. '
                         'N.A. cuenta como tarea resuelta.', color='808080')
    ws.freeze_panes = 'A5'
    for col, ancho in zip('ABCDEGJKM', [8, 13, 7, 11, 9, 11, 10, 9, 12]):
        ws.column_dimensions[col].width = ancho


def slug(t):
    t = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    return re.sub(r'[^A-Za-z0-9]+', '', t) or 'Planta'


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('fichero', help='plantas/<planta>.json o plantas.csv de la toolbox')
    ap.add_argument('-o', '--salida', help='xlsx de salida (default Seguimiento_PEM_<planta>.xlsx)')
    ap.add_argument('--sin-ejemplo', action='store_true',
                    help='no rellenar la fila de ejemplo en la primera TCU')
    args = ap.parse_args()

    ruta = Path(args.fichero)
    planta, ncus = cargar_topologia(ruta)
    wb = Workbook()
    wb.active.title = 'Resumen'
    ult_por_ncu = {}
    for n, d in ncus.items():
        ult_por_ncu[n] = hoja_ncu(wb, n, d, con_ejemplo=not args.sin_ejemplo)
    hoja_resumen(wb, planta, ncus, ult_por_ncu)

    salida = Path(args.salida) if args.salida else Path(f'Seguimiento_PEM_{slug(planta)}.xlsx')
    wb.save(salida)
    tot_tcus = sum(fin - ini + 1 for d in ncus.values() for _, ini, fin in d['gws'])
    print(f'{salida}: {len(ncus)} NCUs, {tot_tcus} TCUs ({planta})')


if __name__ == '__main__':
    sys.exit(main())
