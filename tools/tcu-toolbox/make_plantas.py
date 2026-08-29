#!/usr/bin/env python3
"""Genera plantas.json para TCU Toolbox a partir del config del SCADA.

Lee ``config/plants.yml`` (el mismo fichero que usa el collector) y emite un
``plantas.json`` con una entrada por NCU y puerto de gateway. La toolbox habla
con los TCU a traves del passthrough Modbus de la NCU (unit id = numero de
TCU), que en El Burgo escucha en los puertos 503/504 (uno por gateway Zigbee);
por eso se generan entradas por puerto y no una sola por NCU.

Fuente de verdad de los gateways: la clave opcional ``gateways`` de cada NCU
en plants.yml (el collector la ignora)::

    gateways:
      - { puerto: 503, tcu_ini: 1,  tcu_fin: 56 }
      - { puerto: 504, tcu_ini: 57, tcu_fin: 108 }

Con ``gateways`` declarado, el fichero generado no necesita retoques y puede
regenerarse automaticamente (hay un workflow de GitHub Actions que lo hace en
cada cambio de plants.yml). Si una NCU no lo declara, se cae al modo antiguo:
una entrada por puerto de --puertos con rango 1..tcu_count, a ajustar a mano.

Uso (desde tools/tcu-toolbox/):

    python make_plantas.py                          # usa ../../config/plants.yml
    python make_plantas.py --plants otra/plants.yml
    python make_plantas.py --puertos 503 504        # fallback si no hay gateways
    python make_plantas.py --salida plantas.json

Solo necesita PyYAML (ya esta en requirements del collector).
"""

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path


# Plantas que ya tienen fichero con otro nombre, del modo plants.yml. Ver `destino`, abajo.
MISMO_FICHERO = {"23003": "elburgo.json"}


def rsus_de_celda(texto):
    """Los NUMEROS de estacion de una celda 'RSU': '8' -> [8]; '8\n9' -> [8, 9]; '-' -> [].

    A NIVEL DE MODULO A PROPOSITO, para que `test_columnas_rsu.py` pruebe ESTA y no una copia.
    Tenerla copiada en el banco ya se quedo atras una vez: el banco daba verde con la version
    vieja mientras el codigo real hacia otra cosa, que es la peor manera de tener un banco.

    Separadores: salto de linea, coma, punto y coma, barra y espacio. El salto de linea es el
    que importa —la NCU15 de Ayora trae sus dos, la 8 y la 9, en dos lineas de la misma celda—
    y el `.0` se quita POR TROZO, que Excel lo mete en cada numero y no solo al final.
    """
    trozos = [re.sub(r"\.0$", "", t.strip()) for t in re.split(r"[\n\r,;/ ]+", str(texto or ""))]
    return [int(x) for x in trozos if re.match(r"^\d+$", x)]


def parse_rangos(texto):
    """Los TRAMOS de una celda de 'Esclavos': [(ini, fin), ...], vacio si no hay ninguno.

    UNA NCU NO SIEMPRE TIENE UN SOLO TRAMO. En la hoja, la celda de la NCU 16 de San Jose
    trae CINCO, uno por linea:

        15-19
        27-35
        53-61
        68-76
        95-103

    Esto entendia un solo rango y devolvia None para esas celdas, y sin rango la NCU se
    caia ENTERA del fichero, en silencio. Es el bug que dejo a San Jose con 16 NCU de 21 y
    con tres HSU sin poder asignar: las cinco que faltaban -7, 12, 16, 17 y 19- son
    exactamente las cinco con varios tramos. Ya estaba arreglado en `ips.html`; aqui no.

    Se parte por saltos de linea, comas, puntos y coma Y ESPACIOS. Lo de los espacios
    es el mismo bug otra vez, un separador mas tarde: la tabla de topologia guarda esa
    celda en una linea -«1-13 15-23»- y aqui se partia solo por lineas y comas, asi que
    la celda entera no casaba con nada, devolvia [] y la NCU volvia a caerse ENTERA del
    fichero en silencio. Le paso a la NCU7 de Ayora. Los espacios se quitan antes de
    los guiones, para que «1 - 13» siga siendo un tramo y no los TCU 1 y 13.
    """
    out = []
    for trozo in re.split(r"[\n\r,;\s]+", re.sub(r"\s*-\s*", "-", str(texto or ""))):
        t = trozo.strip()
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", t)
        if m:
            out.append((int(m.group(1)), int(m.group(2))))
        elif re.match(r"^\d+$", t):
            out.append((int(t), int(t)))
    return out


def parse_rango(texto):
    """El PRIMER tramo de la celda, para quien solo pueda con uno. '-'/vacio -> None."""
    r = parse_rangos(texto)
    return r[0] if r else None


def slug(texto):
    t = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", t.lower()).strip("-") or "planta"



def revisa_gateways(entradas, quien):
    """Dos cosas que la hoja no puede decir por si sola y se ven juntando sus rangos.

    EL SOLAPE NUNCA ES LEGITIMO. Un TCU cuelga de UN gateway. Si sale en los dos,
    la toolbox lo sondea por los dos puertos y lo da por caido cuando no contesta
    por el que no es el suyo. Paso en San Jose NCU3 con el TCU 46 y no se veia
    mirando la hoja: hay que juntar los dos rangos para que salte.

    EL HUECO, EN CAMBIO, PUEDE SER BUENO. Un TCU retirado deja su numero vacio y
    los demas no se renumeran -Ayora NCU7 tiene el 14 asi, y El Burgo el 108-. Por
    eso esto solo lo CUENTA: es para mirarlo, no para corregirlo a ciegas. Pero
    tambien fue el sintoma de San Jose NCU9, donde el 49 no colgaba de ninguno
    porque el GW1 acababa en 48 y el GW2 empezaba en 50.
    """
    porgw = {}
    for e in entradas:
        m = re.search(r"NCU\s*(\d+)", e.get("nombre") or "")
        g = re.search(r"GW\s*(\d+)", e.get("nombre") or "")
        if not m or e.get("tcu_ini") is None:
            continue
        porgw.setdefault((int(m.group(1)), int(g.group(1)) if g else 1), set()).update(
            range(int(e["tcu_ini"]), int(e["tcu_fin"]) + 1))
    ncus = {}
    for (n, g), s in porgw.items():
        ncus.setdefault(n, {})[g] = s
    solapes, huecos = [], []
    for n, gs in sorted(ncus.items()):
        for a in sorted(gs):
            for b in sorted(gs):
                if a < b and (gs[a] & gs[b]):
                    solapes.append((n, a, b, sorted(gs[a] & gs[b])))
        todos = set().union(*gs.values())
        falta = sorted(set(range(min(todos), max(todos) + 1)) - todos)
        if falta:
            huecos.append((n, falta))
    for n, a, b, tcus in solapes:
        print("  !! %s NCU%d: el TCU %s cuelga del GW%d Y del GW%d. Un TCU es de UN gateway: "
              "asi se sondea por los dos puertos y sale caido por el que no es el suyo."
              % (quien, n, tcus if len(tcus) > 1 else tcus[0], a, b))
    for n, falta in huecos:
        print("  %s NCU%d: los TCU %s no cuelgan de ningun gateway. Puede ser bueno (TCU "
              "retirada, que no renumera a las demas) o ser un rango mal puesto."
              % (quien, n, falta if len(falta) <= 12 else falta[:12] + ["..."]))
    return solapes, huecos


def modo_excel(ruta, hoja, puertos, excluir, comentario_extra):
    """Lee la hoja 'Direcciones IP' del Excel maestro y devuelve
    {(num, proyecto): [entradas plantas.json]}. Solo usa IP NCU y los rangos
    de esclavos por gateway: las columnas de credenciales NI SE LEEN."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Falta openpyxl: pip install openpyxl")
    wb = load_workbook(ruta, data_only=True)
    if hoja not in wb.sheetnames:
        sys.exit(f"El Excel no tiene la hoja '{hoja}' (tiene: {wb.sheetnames})")
    filas = list(wb[hoja].iter_rows(values_only=True))
    head = [str(c).strip() if c is not None else "" for c in filas[0]]

    def col(nombre, desde=0):
        try:
            return head.index(nombre, desde)
        except ValueError:
            sys.exit(f"La hoja '{hoja}' no tiene la columna '{nombre}'")

    def col_opcional(*nombres):
        """Igual que col() pero devuelve None si no esta: para columnas que
        todavia no existen en todos los Excel."""
        for n in nombres:
            if n in head:
                return head.index(n)
        return None

    i_num, i_proy, i_ncu = col("Nº"), col("Proyecto"), col("NCU")
    i_ip = col("IP NCU")
    i_esc1 = col("Esclavos")
    i_esc2 = col("Esclavos", col("IP GW 2"))   # segundo 'Esclavos', tras IP GW 2
    # LA IP DE CADA GATEWAY. Esta en la hoja desde siempre y no se leia: 'IP GW 1'
    # e 'IP GW 2' solo se usaban como MARCADOR de columna para encontrar el
    # segundo 'Esclavos' y las RSU. Sin su valor, el JSON solo lleva la IP de la
    # NCU —que es su MODBUS TCP, 503/504— y quien tiene que hablar con el gateway
    # (el ConnectPort DIGI, por HTTP/RCI y telnet) se queda sin direccion: la pone
    # a mano o, peor, usa la del Modbus y mide contra el sitio equivocado.
    i_gw1 = col_opcional("IP GW 1", "IP GW1", "IP GW")
    i_gw2 = col_opcional("IP GW 2", "IP GW2")
    if i_gw1 is None:
        print(f"aviso: la hoja '{hoja}' no tiene columna 'IP GW 1'; el JSON saldra sin "
              "la IP de los gateways y habra que ponerla a mano para medir cobertura")
    # El numero de esclavo Modbus de la HSU (estacion meteo) de cada NCU. La
    # hoja no lo trae todavia: en cuanto exista una columna 'HSU' (o
    # 'HSU esclavo'), con el numero por fila de NCU, sale solo en el JSON y la
    # toolbox lo preselecciona en vez de tirar del 185 por defecto.
    # Admite un numero ('230') o varios si esa NCU lleva mas de una estacion
    # ('230,231'), en el orden de los huecos HSU1, HSU2... de la cache de la NCU.
    i_hsu = col_opcional("HSU esclavo", "Esclavo HSU", "HSU")
    if i_hsu is None:
        print(f"aviso: la hoja '{hoja}' no tiene columna 'HSU esclavo'; se conservan "
              "los que ya tenga el JSON y el resto saldra sin el (la toolbox usara 185)")
    # QUE estaciones lleva cada NCU, si lo dicen las columnas 'RSU' (una por gateway).
    # En esa celda va el ORDEN DE LA ESTACION dentro de la planta (1, 2, 3...), NO el
    # esclavo Modbus: volcarlo en hsu_esclavo mandaria a la toolbox a hablar con el
    # esclavo 5, que es un TCU. Pero tampoco vale con contarlo, que es lo que se hacia:
    # ese numero es lo que dice QUE HSU cuelga de QUE NCU, y sin el hay que deducirlo
    # por cercania. Va a `rsu`, que es el campo que ya lee la toolbox.
    # LA COLUMNA 'RSU' DEL GW1, y OJO CON LA CABECERA DE LA IZQUIERDA. La hoja abre con un
    # bloque de totales por planta —'NCUs', 'RSUs', 'TCUs'— y 'RSUs' TAMBIEN empieza por RSU.
    # Buscar por prefijo desde el principio agarra ESA: San Jose tiene 8 RSU en total, asi que
    # su primera NCU salia con rsu [8], que es el numero de la estacion de la NCU 21. Se busca
    # solo dentro del bloque del gateway, o sea a partir de 'IP GW 1', y se descarta el plural.
    def col_rsu(desde):
        for k in range(desde, len(head)):
            t = str(head[k] or "").strip().upper()
            if t.startswith("RSU") and t != "RSUS":
                return k
        return None

    _gw1 = head.index("IP GW 1") if "IP GW 1" in head else 0
    i_rsu1 = col_rsu(_gw1)
    i_rsu2 = None
    if i_rsu1 is not None and "IP GW 2" in head:
        # LA SEGUNDA COLUMNA 'RSU' NO SIEMPRE SE LLAMA IGUAL. Buscar el texto exacto
        # 'RSU' falla si la hoja pone 'RSU 2', 'RSU GW2' o le sobra un espacio, y el
        # fallo es SILENCIOSO: se cuentan solo las del GW1 y nadie se entera. El
        # sintoma esta a la vista en El Burgo, que tiene DOS estaciones por NCU -una
        # por gateway, comprobado en campo- y sale con hsus 1.
        i_rsu2 = col_rsu(head.index("IP GW 2"))
        if i_rsu2 is None:
            print(f"aviso: la hoja '{hoja}' no tiene una segunda columna 'RSU' despues de "
                  "'IP GW 2'; solo se contaran las estaciones del GW1")

    plantas = {}
    actual = None
    ncu_auto = 0
    ultima_entrada = None      # las entradas de la ultima NCU con IP, para las filas de continuacion
    for fila in filas[2:]:
        def v(i):
            return str(fila[i]).strip() if (i < len(fila) and fila[i] is not None) else ""
        num, proy = v(i_num), v(i_proy)
        if num and proy:
            num = re.sub(r"\.0$", "", num)
            actual = (num, proy)
            ncu_auto = 0
            plantas.setdefault(actual, [])
        if not actual:
            continue
        # El Excel solo pone el nombre en la PRIMERA fila de cada planta, asi
        # que usar el de la fila dejaba " NCU2", " NCU3"... sin nombre. Y la
        # toolbox agrupa la entrada "(Planta completa)" por ese prefijo: con
        # los nombres partidos salian dos grupos, uno con la NCU1 sola (que se
        # descarta por tener menos de dos) y otro con el resto. Ayora se
        # quedaba con 15 de 16 NCUs.
        proy = actual[1]

        def rsus_por_gw():
            """QUE RSU declara esta fila EN CADA GATEWAY: [[n del gw1], [n del gw2]].

               Dos cosas que la hoja dice y se tiraban:

               EL GATEWAY, por la COLUMNA. Hay una 'RSU' antes de 'IP GW 2' y otra despues,
               exactamente igual que las dos de 'Esclavos'. Se leian las dos desde siempre
               —`i_rsu1` e `i_rsu2`— pero solo para SUMARLAS.

               Y EL NUMERO, que es lo gordo. En esa celda va el ORDEN DE LA ESTACION dentro
               de la planta: un 5 ahi significa «la HSU 5 cuelga de esta NCU». Se estaba
               comprobando con un regex y tirando: solo se contaba. Con el numero, la hoja
               dice cada HSU con su NCU y su gateway sin deducir nada — es justo lo que
               Ayora ya tiene como `rsu: [8, 9]` y lo que a San Jose le falta."""
            return [rsus_de_celda(v(i)) if i is not None else [] for i in (i_rsu1, i_rsu2)]

        def rsus_fila():
            """Cuantas RSU/HSU declara esta fila, entre los dos gateways."""
            return sum(len(x) for x in rsus_por_gw())

        def esclavos_fila():
            """Los esclavos Modbus que declara esta fila: '230' o '230,231'."""
            if i_hsu is None:
                return []
            crudo = re.sub(r"\.0$", "", v(i_hsu))
            return [int(x) for x in re.split(r"[,;/ ]+", crudo) if re.match(r"^\d+$", x)]

        ip = v(i_ip)
        if not re.match(r"^\d+\.\d+\.\d+\.\d+$", ip):
            # Una NCU con DOS estaciones ocupa dos filas: la segunda solo lleva
            # el numero de RSU, sin NCU ni IP, y cuelga de la NCU de arriba.
            # Ayora es asi: la NCU15 tiene la 8 y la 9. Saltandola entera se
            # perdia una de las diez.
            if ultima_entrada is not None and rsus_fila():
                porgw = rsus_por_gw()
                for e in ultima_entrada:
                    e["hsus"] = e.get("hsus", 0) + rsus_fila()
                    # `hsus_gw` y `rsu` SIN mezclar los gateways: la fila de continuacion
                    # trae su RSU en una de las dos columnas, igual que cualquier otra, asi
                    # que va a la del gateway que le toca. Asi la NCU15 de Ayora acaba con
                    # rsu [8, 9] en su GW1, que es donde estan las dos.
                    k = e.get("_gwidx")
                    if k is not None and porgw[k]:
                        e["hsus_gw"] = e.get("hsus_gw", 0) + len(porgw[k])
                        e["rsu"] = e.get("rsu", []) + porgw[k]
                    for esc in esclavos_fila():
                        e.setdefault("hsu_esclavos", []).append(esc)
            continue
        ncu_auto += 1
        ntxt = re.sub(r"\.0$", "", v(i_ncu))
        n = int(ntxt) if re.match(r"^\d+$", ntxt) else ncu_auto
        def ip_valida(i):
            x = v(i) if i is not None else ""
            return x if re.match(r"^\d+\.\d+\.\d+\.\d+$", x) else ""

        ips_gw = [ip_valida(i_gw1), ip_valida(i_gw2)]
        rangos = [parse_rangos(v(i_esc1)), parse_rangos(v(i_esc2))]
        # UNA ENTRADA POR TRAMO, no por gateway. La toolbox ya sabe de varias entradas para
        # el mismo gateway -El Burgo tiene la del TCU 109 suelto- y es lo unico que respeta
        # los huecos: un rango 1-103 que en realidad son cinco tramos haria sondear TCU que
        # no existen. Se guarda ademas el indice REAL del gateway (0=GW1, 1=GW2), no la
        # posicion entre los presentes: una NCU con rango solo en el GW2 tendria gidx 1 y su
        # RSU esta en la columna 2.
        gws = [(k, puertos[k], r, j) for k, rr in enumerate(rangos) if k < len(puertos)
               for j, r in enumerate(rr)]
        if not gws:
            # Y QUE NO VUELVA A PASAR EN SILENCIO. Una celda con texto que no produce
            # ningun tramo tira la NCU entera del fichero sin decir nada: asi se
            # perdieron cinco NCUs de San Jose y la 7 de Ayora, y solo se ve cuando
            # alguien va a la planta y le faltan seguidores.
            crudo = " ".join(x for x in (v(i_esc1), v(i_esc2)) if x and x.strip("- "))
            if crudo:
                print(f"aviso: {proy} NCU{re.sub(r'[.]0$', '', v(i_ncu)) or ncu_auto + 1}: "
                      f"la celda de esclavos «{crudo}» no da ningun tramo; esa NCU NO sale "
                      "en el fichero")
            continue
        escl = esclavos_fila()
        nRsu = rsus_fila()
        ultima_entrada = []
        porgw = rsus_por_gw()
        # Cuantos gateways DISTINTOS hay: el sufijo " GW1"/" GW2" depende de eso, no del
        # numero de entradas, que ahora puede ser mayor por los tramos.
        _kgws = sorted({k for k, _p, _r, _j in gws})
        for kgw, puerto, (ini, fin), jtramo in gws:
            sufijo = f" GW{_kgws.index(kgw) + 1}" if len(_kgws) > 1 else ""
            # Y si ese gateway tiene VARIOS tramos, el rango en el nombre para distinguirlos,
            # igual que "El Burgo I NCU2 GW2 (TCU 109-109)" en el modo plants.yml.
            if sum(1 for k, _p, _r, _j in gws if k == kgw) > 1:
                sufijo += f" (TCU {ini}-{fin})"
            entrada = {
                "nombre": f"{proy} NCU{n}{sufijo}",
                "ip": ip,
                # La del ConnectPort DIGI, que NO es la de arriba: `ip` es el
                # Modbus de la NCU y esta es a quien se pregunta por HTTP/RCI y
                # telnet cuando se mide cobertura. `kgw` es el indice REAL del
                # gateway (0=GW1, 1=GW2), que es lo que indexa la columna.
                **({"ip_gw": ips_gw[kgw]} if kgw < len(ips_gw) and ips_gw[kgw] else {}),
                "puerto": puerto,
                "tcu_ini": ini,
                "tcu_fin": fin,
            }
            # La HSU cuelga de UN gateway, no de los dos, y el Excel SI dice cual:
            # lo dice la columna 'RSU' en la que va el numero, que hay una por
            # gateway. Eso es `hsus_gw`, unas lineas mas abajo. Aqui el ESCLAVO se
            # sigue poniendo en las dos entradas porque ese si falta en la hoja, y
            # la toolbox ya avisa de cambiar el puerto si no responde por ese.
            if escl:
                entrada["hsu_esclavos"] = list(escl)
            if nRsu:
                entrada["hsus"] = nRsu
            # LO QUE FALTABA. `hsus` es el total de la NCU repetido en sus dos filas
            # -asi ha sido siempre y asi se queda, que hay quien lo lee- y `hsus_gw`
            # dice cuantas van EN ESTE gateway, que es lo que la columna sabe y se
            # estaba perdiendo. Con eso deja de hacer falta adivinarlo por cercania.
            if porgw[kgw] and jtramo == 0:
                entrada["hsus_gw"] = len(porgw[kgw])
                # QUE estaciones, no solo cuantas. Es el campo que ya lee la toolbox y con
                # el que `meteo_ncu.mjs` empareja cada HSU del DWG con su NCU sin deducir.
                entrada["rsu"] = list(porgw[kgw])
            entrada["_gwidx"] = kgw            # interno, se quita antes de escribir
            plantas[actual].append(entrada)
            ultima_entrada.append(entrada)

    # El indice de gateway era solo para repartir las RSU de las filas de continuacion:
    # fuera del bucle ya no pinta nada y no tiene por que acabar en el JSON.
    for entradas in plantas.values():
        for e in entradas:
            e.pop("_gwidx", None)

    ficheros = []
    for (num, proy), entradas in sorted(plantas.items()):
        if not entradas or num in excluir:
            continue
        # DOS MODOS, UN SOLO FICHERO POR PLANTA. El modo plants.yml nombra por `plant.id`
        # (`elburgo.json`) y este nombra por numero y nombre (`23003-burgo-i.json`), asi que
        # El Burgo acababa DUPLICADO en plantas/ y el portatil de campo lo veria dos veces.
        # El emparejamiento va escrito a la vista porque es una decision, no algo deducible:
        # el Excel llama a esa planta «Burgo I» y plants.yml «El Burgo I», y de ahi no sale.
        destino = Path("plantas") / MISMO_FICHERO.get(str(num), f"{num}-{slug(proy)}.json")
        destino.parent.mkdir(parents=True, exist_ok=True)
        # Los esclavos de las HSUs no estan en el Excel (todavia): si el JSON
        # anterior los traia puestos a mano, regenerar no puede borrarlos. Solo
        # se conservan los que el Excel NO manda; en cuanto la hoja tenga la
        # columna, manda la hoja.
        if destino.exists():
            try:
                previo = json.loads(destino.read_text(encoding="utf-8"))
            except (ValueError, OSError):
                previo = {"plantas": []}

            # NO SE EMPAREJA POR NOMBRE. El Excel llama a la planta «Burgo I» y plants.yml
            # «El Burgo I», asi que emparejar por `nombre` no casaba NINGUNA entrada y la
            # regeneracion se llevaba por delante los esclavos 230/231 de El Burgo, que son
            # dato de campo. La identidad de una entrada es su NCU, su puerto y su rango.
            def clave(e):
                m = re.search(r"NCU\s*(\d+)", str(e.get("nombre") or ""))
                return (m.group(1) if m else "", e.get("puerto"), e.get("tcu_ini"), e.get("tcu_fin"))

            antes = {clave(e): e for e in previo.get("plantas", [])}
            conservados = 0
            for e in entradas:
                viejo = antes.pop(clave(e), None)
                if not viejo:
                    continue
                # REGENERAR SOLO PISA LO QUE LA HOJA DICE. Lo que no dice se queda,
                # sea lo que sea: no hay lista de campos que mantener.
                #
                # La habia -`hsu_esclavos, rsu, hsus_gw, ip_gw`- y se dejaba fuera
                # `repetidores` y `huecos`, que el Excel tampoco trae:
                # la regeneracion de la #222 se llevo por delante los CINCO
                # repetidores de Ayora y los TRES huecos de su NCU7, en silencio.
                # Sin los repetidores no se diagnostican, no entran en el
                # inventario ni en la campana de firmware, y uno con la bateria
                # muerta se lleva por delante lo que cuelga de el. Sin los huecos,
                # tres seguidores que no existen salen OFFLINE en cada barrido y
                # la planta pasa de 751 a 754.
                #
                # La regla nueva no enumera nada: si la entrada recien construida
                # YA trae el campo, manda la hoja; si no lo trae, es que la hoja no
                # opina, y lo que hubiera en el JSON viene de otro sitio -a mano,
                # de los .bat de Sunner, de una medida en campo- y no es suyo para
                # borrarlo. Anadir manana un campo a mano ya no exige acordarse de
                # esta linea, y quitar un dato sigue siendo posible: se quita del
                # JSON, no de la hoja.
                for k, v in viejo.items():
                    if k in e or v is None:
                        continue
                    e[k] = v
                    conservados += 1
            # Y LAS ENTRADAS QUE EL EXCEL NO TRAE TAMPOCO SE TIRAN. En El Burgo hay una fila
            # que no sale de ninguna hoja: el TCU 109 suelto en el GW2 de la NCU2, sacado de
            # los .bat de Sunner. Regenerar desde el Excel la borraba en silencio.
            sobrantes = [e for e in antes.values() if e.get("tcu_ini") is not None]
            if sobrantes:
                entradas.extend(sobrantes)
                entradas.sort(key=lambda e: (clave(e)[0].zfill(3), e.get("puerto") or 0, e.get("tcu_ini") or 0))
                print(f"  {destino}: conservadas {len(sobrantes)} entradas que el Excel no trae "
                      + ", ".join(str(e.get("nombre")) for e in sobrantes))
            if conservados:
                print(f"  {destino}: conservados {conservados} campos que el Excel no trae "
                      "(repetidores, huecos, esclavos de HSU, ip_gw... del JSON anterior)")
        revisa_gateways(entradas, destino.name)
        destino.write_text(json.dumps({
            "_comentario": (f"Planta {num} ({proy}) generada desde el Excel maestro "
                            f"(hoja '{hoja}') por make_plantas.py --excel. Solo topologia: "
                            "sin credenciales. NO subir el Excel al repo." + comentario_extra),
            "version": 1,
            "plantas": entradas,
        }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        ficheros.append((destino, len(entradas)))
        print(f"{destino}: {len(entradas)} entradas ({proy})")
    if not ficheros:
        sys.exit("El Excel no produjo ninguna planta (revisa la hoja y las columnas)")
    return ficheros



# ── modo tarjeta ─────────────────────────────────────────────────────────────
# La tarjeta «IPs de plantas» (factiun-cartera) exporta la topologia de Supabase
# en JSONs con esta misma forma. Trae una cosa que aqui no hay y no se puede
# deducir: la IP del ConnectPort de cada gateway. Este modo la mete y NO TOCA
# NADA MAS: los rangos de TCU los mandan plants.yml y el Excel, no la tarjeta.
#
#   python make_plantas.py --tarjeta IPs.zip
#
# Se le puede dar el .zip tal cual sale del navegador, una carpeta o ficheros
# sueltos.

def _entradas_de(ruta):
    """Los JSON de la tarjeta que haya en `ruta`: .zip, carpeta o fichero."""
    import zipfile
    ruta = Path(ruta)
    if ruta.is_dir():
        return [(f.name, json.loads(f.read_text(encoding="utf-8")))
                for f in sorted(ruta.glob("*.json"))]
    if ruta.suffix.lower() == ".zip":
        with zipfile.ZipFile(ruta) as z:
            return [(n, json.loads(z.read(n).decode("utf-8")))
                    for n in sorted(z.namelist()) if n.lower().endswith(".json")]
    return [(ruta.name, json.loads(ruta.read_text(encoding="utf-8")))]


def _ncu_de(nombre):
    """El NUMERO de NCU del nombre de una entrada. Los prefijos NO casan entre
    fuentes -el Excel dice «Burgo I» y plants.yml «El Burgo I»-, asi que emparejar
    por nombre completo no vale: se empareja por (NCU, puerto)."""
    m = re.search(r"NCU\s*(\d+)", str(nombre or ""))
    return int(m.group(1)) if m else None


def _cobertura(entradas):
    """Que TCUs cubre cada (NCU, puerto), contando los huecos declarados."""
    cob = {}
    for e in entradas:
        n, p = _ncu_de(e.get("nombre")), e.get("puerto")
        if n is None or e.get("tcu_ini") is None:
            continue
        cob.setdefault((n, p), set()).update(
            set(range(int(e["tcu_ini"]), int(e["tcu_fin"]) + 1)) - set(e.get("huecos") or []))
    return cob


def modo_tarjeta(rutas, dir_plantas):
    avisos, escritos, puestas = [], [], 0
    for ruta in rutas:
        for fichero, doc in _entradas_de(ruta):
            entradas = doc.get("plantas") or []
            num = re.match(r"(\d+)", fichero)
            num = num.group(1) if num else ""
            # el mismo emparejamiento de nombres que el modo Excel: si no, El Burgo
            # entra por segunda vez con otro nombre y el portatil de campo lo ve doble
            destino = dir_plantas / MISMO_FICHERO.get(num, fichero)
            if not destino.exists():
                avisos.append(f"{fichero}: no hay {destino}; esta planta no esta en plantas/")
                continue
            doc_local = json.loads(destino.read_text(encoding="utf-8"))
            locales = doc_local.get("plantas") or []
            por_clave = {}
            for e in locales:
                n = _ncu_de(e.get("nombre"))
                if n is not None:
                    por_clave.setdefault((n, e.get("puerto")), []).append(e)

            cambios = 0
            for e in entradas:
                gw = str(e.get("ip_gw") or "").strip()
                if not gw:
                    continue
                if not re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", gw) or any(
                        int(o) > 255 for o in gw.split(".")):
                    avisos.append(f"{fichero}: «{gw}» no es una IP; {e.get('nombre')} se queda sin gateway")
                    continue
                # el error que costaria una jornada de campo: el gateway NUNCA es la
                # IP del Modbus. Si coinciden, la celda esta mal puesta en la tabla.
                if gw == str(e.get("ip") or "").strip():
                    avisos.append(f"{fichero}: {e.get('nombre')} da el gateway igual que el Modbus ({gw}); "
                                  "NO se escribe, revisa la celda de la tabla")
                    continue
                clave = (_ncu_de(e.get("nombre")), e.get("puerto"))
                destinos = por_clave.get(clave)
                if not destinos:
                    avisos.append(f"{fichero}: {e.get('nombre')} (puerto {clave[1]}) no casa con ninguna entrada de {destino.name}")
                    continue
                # un gateway por (NCU, puerto): si ahi hay varios tramos, la IP es la
                # misma para todos (El Burgo NCU2 GW2 y su TCU 109 suelto)
                for d in destinos:
                    if d.get("ip_gw") != gw:
                        d["ip_gw"] = gw
                        cambios += 1

            # Los rangos NO se tocan: se comparan y se cantan. La tarjeta sale de
            # Supabase y plantas/ de plants.yml y del Excel; cuando no coinciden,
            # alguien esta poleando TCUs que no existen o dejandose otras sin medir,
            # y eso lo decide una persona mirando la planta, no este script.
            ct, cl = _cobertura(entradas), _cobertura(locales)
            for k in sorted(set(ct) | set(cl), key=lambda x: (x[0], x[1] or 0)):
                sobra, falta = sorted(cl.get(k, set()) - ct.get(k, set())), sorted(ct.get(k, set()) - cl.get(k, set()))
                if sobra or falta:
                    avisos.append(f"{destino.name}: NCU{k[0]} puerto {k[1]} no cuadra con la tabla"
                                  + (f"; aqui de mas {sobra}" if sobra else "")
                                  + (f"; en la tabla y aqui no {falta}" if falta else ""))
            if cambios:
                destino.write_text(json.dumps(doc_local, ensure_ascii=False, indent=2) + "\n",
                                   encoding="utf-8")
                escritos.append((destino, cambios))
                puestas += cambios
                print(f"{destino}: {cambios} IPs de gateway")

    if not puestas:
        print("Ninguna IP de gateway en el export: es anterior al cambio de la tarjeta, "
              "o la tabla no las tiene puestas para estas plantas.")
    for a in avisos:
        print("  aviso: " + a)
    return escritos, avisos


try:
    import yaml
except ImportError:
    yaml = None


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--plants", default="../../config/plants.yml",
                    help="ruta a plants.yml del SCADA (defecto: ../../config/plants.yml)")
    ap.add_argument("--excel", default=None,
                    help="modo Excel maestro: ruta al .xlsx con la hoja 'Direcciones IP'; "
                         "genera plantas/<num>-<planta>.json por cada planta (sin credenciales)")
    ap.add_argument("--hoja", default="Direcciones IP", help="hoja del Excel (defecto: Direcciones IP)")
    ap.add_argument("--excluir", nargs="*", default=[],
                    help="numeros de proyecto a saltar en modo Excel (p.ej. --excluir 23003)")
    ap.add_argument("--puertos", nargs="+", type=int, default=[503, 504],
                    help="puertos passthrough de la NCU, uno por gateway (defecto: 503 504)")
    ap.add_argument("--tarjeta", nargs="+", default=None,
                    help="mete en plantas/ la IP del ConnectPort de cada gateway desde el export "
                         "de la tarjeta «IPs de plantas» (.zip, carpeta o JSONs). No toca nada mas")
    ap.add_argument("--plantas", default=None,
                    help="carpeta plantas/ a actualizar (defecto: la de al lado de este script)")
    ap.add_argument("--salida", default=None,
                    help="fichero de salida (defecto: plantas/<plant_id>.json, un JSON por planta)")
    args = ap.parse_args()

    if args.tarjeta:
        # por defecto la carpeta de al lado del script, no la del directorio actual:
        # correr esto desde la raiz del repo escribia plantas/ donde no era
        dirp = Path(args.plantas) if args.plantas else Path(__file__).resolve().parent / "plantas"
        modo_tarjeta([Path(r) for r in args.tarjeta], dirp)
        return

    if args.excel:
        modo_excel(Path(args.excel), args.hoja, args.puertos, set(args.excluir), "")
        return

    if yaml is None:
        sys.exit("Falta PyYAML: pip install pyyaml")
    ruta = Path(args.plants)
    if not ruta.exists():
        sys.exit(f"No existe {ruta} (ejecuta desde tools/tcu-toolbox/ o pasa --plants)")

    cfg = yaml.safe_load(ruta.read_text(encoding="utf-8"))
    nombre_planta = (cfg.get("plant") or {}).get("name") or (cfg.get("plant") or {}).get("id") or "Planta"
    ncus = cfg.get("ncus") or []
    if not ncus:
        sys.exit("plants.yml no tiene NCUs")

    plantas = []
    a_mano = False
    for ncu in ncus:
        host = ncu.get("host")
        if not host:
            continue
        gws = ncu.get("gateways") or []
        if gws:
            # mismo puerto = mismo gateway fisico; una fila extra del mismo
            # puerto (p.ej. TCU suelta 109) se distingue por su rango
            indice_gw = {}
            vistos = {}
            for gw in gws:
                puerto = int(gw["puerto"])
                if puerto not in indice_gw:
                    indice_gw[puerto] = len(indice_gw) + 1
                ini = int(gw.get("tcu_ini", 1))
                fin = int(gw.get("tcu_fin", ncu.get("tcu_count") or 1))
                sufijo = f" GW{indice_gw[puerto]}" if len(gws) > 1 else ""
                if puerto in vistos:
                    sufijo += f" (TCU {ini}-{fin})"
                vistos[puerto] = True
                entrada = {
                    "nombre": gw.get("nombre") or f"{nombre_planta} {ncu.get('id', host)}{sufijo}",
                    "ip": host,
                    "puerto": puerto,
                    "tcu_ini": ini,
                    "tcu_fin": fin,
                }
                # La estación meteo cuelga de UN gateway, no de los dos, y ahora
                # el gateway lo dice: `hsu_esclavo` en su fila de plants.yml
                # (230 en el GW1, 231 en el GW2). Antes esto no se sabía y el
                # esclavo se repetía en las entradas de la NCU entera, dejando a
                # la toolbox adivinando el puerto.
                esc = gw.get("hsu_esclavo", gw.get("hsu_esclavos"))
                if esc is not None:
                    escl = [int(x) for x in (esc if isinstance(esc, (list, tuple)) else [esc])]
                    if escl:
                        entrada["hsus"] = len(escl)
                        entrada["hsu_esclavos"] = escl
                plantas.append(entrada)
        else:
            # fallback sin gateways declarados: rangos 1..tcu_count a ajustar a mano
            a_mano = True
            tcus = int(ncu.get("tcu_count") or 0) or 1
            for i, puerto in enumerate(args.puertos, start=1):
                sufijo = f" GW{i}" if len(args.puertos) > 1 else ""
                plantas.append({
                    "nombre": f"{nombre_planta} {ncu.get('id', host)}{sufijo}",
                    "ip": host,
                    "puerto": puerto,
                    "tcu_ini": 1,
                    "tcu_fin": tcus,
                })

    comentario = f"Generado desde {ruta.name} por make_plantas.py. NO editar a mano: declara los gateways en plants.yml y regenera."
    if a_mano:
        comentario = (f"Generado desde {ruta.name} por make_plantas.py. Hay NCUs sin 'gateways' en plants.yml: "
                      "ajusta sus rangos tcu_ini/tcu_fin a mano o, mejor, declaralos en plants.yml y regenera.")
    salida = {
        "_comentario": comentario,
        "version": 1,
        "plantas": plantas,
    }
    if args.salida:
        destino = Path(args.salida)
    else:
        pid = (cfg.get("plant") or {}).get("id") or "planta"
        destino = Path("plantas") / f"{pid}.json"
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(salida, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"{destino}: {len(plantas)} entradas ({len(ncus)} NCUs)")


if __name__ == "__main__":
    main()
